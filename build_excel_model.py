"""
Generate the Excel Equity Valuation Model (.xlsm)
Includes: Assumptions, Income Projection, FCF, DCF, Multiples, Sensitivity, Dashboard
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ── Color Constants ────────────────────────────────────────────────────────────
DARK_NAVY = "0D1B2A"
MID_BLUE  = "1E6FBF"
ACCENT    = "F0A500"
LIGHT_BG  = "F5F7FA"
WHITE     = "FFFFFF"
INPUT_BLUE = "0000FF"   # Hardcoded inputs
FORMULA_BLK = "000000"  # Formulas
LINK_GREEN  = "008000"  # Cross-sheet links

# ── Style Helpers ──────────────────────────────────────────────────────────────

def header_font(size=10, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def input_font(size=9):
    return Font(name="Arial", size=9, bold=False, color=INPUT_BLUE)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def right():
    return Alignment(horizontal="right", vertical="center")

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col+i, value=h)
        cell.font = header_font()
        cell.fill = fill(MID_BLUE)
        cell.alignment = center()
        cell.border = thin_border()

def write_section_title(ws, row, col, title):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = Font(name="Arial", size=11, bold=True, color=DARK_NAVY)
    cell.fill = fill(LIGHT_BG)


# ── Sheet 1: Assumptions ───────────────────────────────────────────────────────

def build_assumptions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26

    # Title banner
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = "EQUITY VALUATION MODEL — ASSUMPTIONS"
    cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    cell.fill = fill(DARK_NAVY)
    cell.alignment = center()

    ws["A2"] = "Color Code: Blue = Input  |  Black = Formula  |  Green = Cross-sheet link"
    ws["A2"].font = Font(name="Arial", size=8, color="666666", italic=True)

    rows = [
        (4, "COMPANY INFORMATION", None, None),
        (5, "Company Name", "DEMO Corp.", "Enter full company name"),
        (6, "Ticker Symbol", "DEMO", "Exchange:Ticker"),
        (7, "Reporting Currency", "USD ($M)", "All figures in $M unless noted"),
        (8, "Valuation Date", "2024-12-31", "YYYY-MM-DD"),

        (10, "CAPITAL STRUCTURE (WACC INPUTS)", None, None),
        (11, "Risk-Free Rate (10Y Treasury)", 0.040, "Source: US Treasury"),
        (12, "Equity Risk Premium", 0.055, "Source: Damodaran (2024)"),
        (13, "Levered Beta", 1.15, "Source: Bloomberg LTM"),
        (14, "Pre-Tax Cost of Debt", 0.045, "YTM of LT debt"),
        (15, "Effective Tax Rate", 0.210, "Blended effective rate"),
        (16, "Debt Weight (D/V)", 0.250, "Market value weights"),
        (17, "Equity Weight (E/V)", "=1-B16", "Auto-calculated"),
        (18, "Size Premium", 0.010, "Duff & Phelps"),
        (19, "Specific Risk Premium", 0.000, "Company-specific adj."),

        (21, "DCF ASSUMPTIONS", None, None),
        (22, "Projection Horizon (Years)", 5, ""),
        (23, "LTM Revenue ($M)", 12600, "Most recent 12 months"),
        (24, "Terminal Growth Rate (g)", 0.025, "Long-run nominal growth"),
        (25, "Exit Multiple (EV/EBITDA)", 12.0, "Alt. TV method"),
        (26, "Terminal Value Method", "Gordon Growth", "Gordon Growth or Exit Multiple"),

        (28, "FCF DRIVER ASSUMPTIONS", None, None),
        (29, "Revenue Growth — Year 1", 0.120, ""),
        (30, "Revenue Growth — Year 2", 0.100, ""),
        (31, "Revenue Growth — Year 3", 0.090, ""),
        (32, "Revenue Growth — Year 4", 0.080, ""),
        (33, "Revenue Growth — Year 5", 0.070, ""),
        (34, "EBITDA Margin (stable)", 0.270, "% of Revenue"),
        (35, "D&A (% of Revenue)", 0.050, ""),
        (36, "CapEx (% of Revenue)", 0.060, ""),
        (37, "Δ NWC (% of Revenue Δ)", 0.015, ""),

        (39, "BALANCE SHEET (LTM)", None, None),
        (40, "Total Debt ($M)", 4200, "Gross financial debt"),
        (41, "Cash & Equivalents ($M)", 2100, ""),
        (42, "Net Debt ($M)", "=B40-B41", "Auto-calculated"),
        (43, "Shares Outstanding (M)", 500, "Fully diluted"),
    ]

    for row_num, label, value, note in rows:
        ws.row_dimensions[row_num].height = 18
        a = ws.cell(row=row_num, column=1, value=label)
        if value is None:
            # Section header
            a.font = Font(name="Arial", size=9, bold=True, color=DARK_NAVY)
            a.fill = fill(LIGHT_BG)
            ws.merge_cells(f"A{row_num}:C{row_num}")
        else:
            a.font = body_font()
            b = ws.cell(row=row_num, column=2, value=value)
            if isinstance(value, str) and value.startswith("="):
                b.font = body_font(color=LINK_GREEN)
            elif isinstance(value, float) and value < 1:
                b.font = input_font()
                b.number_format = "0.0%"
            elif isinstance(value, (int, float)):
                b.font = input_font()
                b.number_format = '#,##0.0'
            else:
                b.font = input_font()
            b.alignment = right()
            b.border = thin_border()

            c = ws.cell(row=row_num, column=3, value=note)
            c.font = Font(name="Arial", size=8, color="888888", italic=True)

        a.border = thin_border()


# ── Sheet 2: Income Statement Projection ──────────────────────────────────────

def build_income_projection(ws, assump_sheet="Assumptions"):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    for col in ["B","C","D","E","F"]:
        ws.column_dimensions[col].width = 14

    # Banner
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")
    ws["A1"].value = "INCOME STATEMENT PROJECTION"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = fill(DARK_NAVY)
    ws["A1"].alignment = center()

    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    write_header_row(ws, 3, headers)

    year_row = 4
    for i, y in enumerate(["FY+1","FY+2","FY+3","FY+4","FY+5"], start=2):
        ws.cell(row=year_row, column=i, value=y).font = body_font(bold=True)

    # Assumptions references (col B = Year1 = col B in this sheet)
    growth_refs = [f"Assumptions!B{29+i}" for i in range(5)]  # B29..B33

    line_items = [
        ("Revenue ($M)", [
            f"=Assumptions!B23*(1+{growth_refs[0]})",
            "=B6*(1+Assumptions!B30)",
            "=C6*(1+Assumptions!B31)",
            "=D6*(1+Assumptions!B32)",
            "=E6*(1+Assumptions!B33)",
        ], FORMULA_BLK, '#,##0.0'),
        ("Revenue Growth", [
            f"={growth_refs[0]}",
            "=C6/B6-1","=D6/C6-1","=E6/D6-1","=F6/E6-1"
        ], FORMULA_BLK, "0.0%"),
        ("EBITDA ($M)", [
            "=B6*Assumptions!B34","=C6*Assumptions!B34",
            "=D6*Assumptions!B34","=E6*Assumptions!B34","=F6*Assumptions!B34"
        ], FORMULA_BLK, '#,##0.0'),
        ("EBITDA Margin", [
            "=B8/B6","=C8/C6","=D8/D6","=E8/E6","=F8/F6"
        ], FORMULA_BLK, "0.0%"),
        ("D&A ($M)", [
            "=B6*Assumptions!B35","=C6*Assumptions!B35",
            "=D6*Assumptions!B35","=E6*Assumptions!B35","=F6*Assumptions!B35"
        ], FORMULA_BLK, '#,##0.0'),
        ("EBIT ($M)", [
            "=B8-B10","=C8-C10","=D8-D10","=E8-E10","=F8-F10"
        ], FORMULA_BLK, '#,##0.0'),
        ("NOPAT = EBIT × (1-T)", [
            "=B12*(1-Assumptions!B15)","=C12*(1-Assumptions!B15)",
            "=D12*(1-Assumptions!B15)","=E12*(1-Assumptions!B15)","=F12*(1-Assumptions!B15)",
        ], FORMULA_BLK, '#,##0.0'),
    ]

    for r, (label, formulas, color, fmt) in enumerate(line_items, start=6):
        ws.cell(row=r, column=1, value=label).font = body_font()
        ws.cell(row=r, column=1).border = thin_border()
        for c, formula in enumerate(formulas, start=2):
            cell = ws.cell(row=r, column=c, value=formula)
            cell.font = Font(name="Arial", size=9, color=color)
            cell.number_format = fmt
            cell.alignment = right()
            cell.border = thin_border()


# ── Sheet 3: FCF & DCF ────────────────────────────────────────────────────────

def build_dcf(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 14

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:G1")
    ws["A1"].value = "DCF VALUATION — FREE CASH FLOW & ENTERPRISE VALUE"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = fill(DARK_NAVY)
    ws["A1"].alignment = center()

    # WACC section
    write_section_title(ws, 3, 1, "WACC Calculation")
    wacc_rows = [
        (4, "Cost of Equity (CAPM)", "=Assumptions!B11+Assumptions!B13*Assumptions!B12+Assumptions!B18+Assumptions!B19", "0.0%"),
        (5, "After-Tax Cost of Debt", "=Assumptions!B14*(1-Assumptions!B15)", "0.0%"),
        (6, "WACC", "=B4*Assumptions!B17+B5*Assumptions!B16", "0.00%"),
    ]
    for row, label, formula, fmt in wacc_rows:
        ws.cell(row=row, column=1, value=label).font = body_font()
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = Font(name="Arial", size=9, bold=(row==6), color=FORMULA_BLK)
        cell.number_format = fmt
        cell.border = thin_border()

    # FCF Bridge
    write_section_title(ws, 8, 1, "Free Cash Flow Bridge")
    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    write_header_row(ws, 9, headers)

    fcf_rows = [
        ("NOPAT ($M)", ["='Income Projection'!B13","='Income Projection'!C13",
                         "='Income Projection'!D13","='Income Projection'!E13","='Income Projection'!F13"]),
        ("+ D&A ($M)",  ["='Income Projection'!B10","='Income Projection'!C10",
                          "='Income Projection'!D10","='Income Projection'!E10","='Income Projection'!F10"]),
        ("- CapEx ($M)", [f"=-'Income Projection'!B6*Assumptions!B36",
                          f"=-'Income Projection'!C6*Assumptions!B36",
                          f"=-'Income Projection'!D6*Assumptions!B36",
                          f"=-'Income Projection'!E6*Assumptions!B36",
                          f"=-'Income Projection'!F6*Assumptions!B36"]),
        ("- Δ NWC ($M)", ["=-('Income Projection'!B6-Assumptions!B23)*Assumptions!B37",
                           "=-('Income Projection'!C6-'Income Projection'!B6)*Assumptions!B37",
                           "=-('Income Projection'!D6-'Income Projection'!C6)*Assumptions!B37",
                           "=-('Income Projection'!E6-'Income Projection'!D6)*Assumptions!B37",
                           "=-('Income Projection'!F6-'Income Projection'!E6)*Assumptions!B37"]),
        ("= Free Cash Flow ($M)", ["=B10+B11+B12+B13","=C10+C11+C12+C13",
                                    "=D10+D11+D12+D13","=E10+E11+E12+E13","=F10+F11+F12+F13"]),
        ("Discount Factor", ["=1/(1+$B$6)^1","=1/(1+$B$6)^2","=1/(1+$B$6)^3","=1/(1+$B$6)^4","=1/(1+$B$6)^5"]),
        ("PV of FCF ($M)", ["=B14*B15","=C14*C15","=D14*D15","=E14*E15","=F14*F15"]),
    ]

    for r, (label, formulas) in enumerate(fcf_rows, start=10):
        ws.cell(row=r, column=1, value=label).font = body_font(bold=(r==14))
        ws.cell(row=r, column=1).border = thin_border()
        for c, f in enumerate(formulas, start=2):
            cell = ws.cell(row=r, column=c, value=f)
            cell.font = Font(name="Arial", size=9, color=LINK_GREEN if "Income" in f or "Assumptions" in f else FORMULA_BLK)
            cell.number_format = "0.000" if r==15 else "#,##0.0"
            cell.alignment = right()
            cell.border = thin_border()

    # Valuation Summary
    write_section_title(ws, 19, 1, "Enterprise Value Bridge")
    val_rows = [
        (20, "Sum of PV(FCFs) ($M)", "=SUM(B16:F16)"),
        (21, "Terminal Value ($M)", "='Income Projection'!F8/(DCF!B6-Assumptions!B24)*(1+Assumptions!B24)"),
        (22, "PV(Terminal Value) ($M)", "=B21/(1+B6)^5"),
        (23, "Enterprise Value ($M)", "=B20+B22"),
        (24, "(-) Net Debt ($M)", "=Assumptions!B42"),
        (25, "Equity Value ($M)", "=B23-B24"),
        (26, "Shares Outstanding (M)", "=Assumptions!B43"),
        (27, "Intrinsic Value per Share", "=B25/B26"),
        (28, "TV as % of EV", "=B22/B23"),
    ]
    for row, label, formula in val_rows:
        ws.cell(row=row, column=1, value=label).font = body_font(bold=(row in [23,25,27]))
        ws.cell(row=row, column=1).border = thin_border()
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = Font(name="Arial", size=9, bold=(row in [23,25,27]),
                          color=FORMULA_BLK)
        if row == 28:
            cell.number_format = "0.0%"
        elif row == 27:
            cell.number_format = "$#,##0.00"
        else:
            cell.number_format = "#,##0.0"
        cell.alignment = right()
        cell.border = thin_border()
        if row in [23, 25, 27]:
            ws.cell(row=row, column=1).fill = fill(LIGHT_BG)
            cell.fill = fill(LIGHT_BG)


# ── Sheet 4: Sensitivity ───────────────────────────────────────────────────────

def build_sensitivity(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:H1")
    ws["A1"].value = "SENSITIVITY ANALYSIS — EQUITY VALUE ($M)"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = fill(DARK_NAVY)
    ws["A1"].alignment = center()

    ws["A3"] = "Equity Value by WACC × Terminal Growth Rate"
    ws["A3"].font = Font(name="Arial", size=10, bold=True, color=DARK_NAVY)

    # Column headers = terminal growth rates
    growth_rates = [0.015, 0.020, 0.025, 0.030, 0.035]
    wacc_rates   = [0.070, 0.075, 0.080, 0.085, 0.090, 0.095, 0.100]

    ws.cell(row=4, column=1, value="WACC \\ g").font = header_font()
    ws.cell(row=4, column=1).fill = fill(MID_BLUE)

    for col_i, g in enumerate(growth_rates, start=2):
        c = ws.cell(row=4, column=col_i, value=g)
        c.font = header_font()
        c.fill = fill(MID_BLUE)
        c.number_format = "0.0%"
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    for row_i, w in enumerate(wacc_rates, start=5):
        rc = ws.cell(row=row_i, column=1, value=w)
        rc.font = header_font()
        rc.fill = fill(MID_BLUE)
        rc.number_format = "0.0%"
        rc.alignment = center()

        for col_i, g in enumerate(growth_rates, start=2):
            # Simplified formula: PV FCFs + TV
            # We use a static approximation referencing Assumptions
            formula = (
                f"=DCF!B20 + "
                f"'Income Projection'!F8*(1+{g})/({w}-{g})/(1+{w})^5 "
                f"-Assumptions!B42"
            )
            cell = ws.cell(row=row_i, column=col_i, value=formula)
            cell.font = Font(name="Arial", size=9, color=FORMULA_BLK)
            cell.number_format = "#,##0"
            cell.alignment = right()
            cell.border = thin_border()

            # Highlight base case
            if abs(w - 0.085) < 0.001 and abs(g - 0.025) < 0.001:
                cell.fill = fill(ACCENT)
                cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)


# ── Sheet 5: Comparable Companies ─────────────────────────────────────────────

def build_comps(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E","F","G","H","I"]:
        ws.column_dimensions[col].width = 14

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:I1")
    ws["A1"].value = "COMPARABLE COMPANY ANALYSIS — TRADING MULTIPLES"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = fill(DARK_NAVY)
    ws["A1"].alignment = center()

    headers = ["Company", "Market Cap", "Net Debt", "EV", "Revenue",
               "EBITDA", "EV/Revenue", "EV/EBITDA", "P/E"]
    write_header_row(ws, 3, headers)
    for c in range(1, 10):
        ws.cell(row=3, column=c).alignment = center()

    comps_data = [
        ("Alpha Tech",   42000, 2500, None, 12600, 3150, None, None, None),
        ("Beta Systems", 28000, 1200, None, 7800,  1872, None, None, None),
        ("Gamma Corp",   65000, 4100, None, 18200, 4914, None, None, None),
        ("Delta Inc",    19500,  800, None, 5400,  1188, None, None, None),
        ("Epsilon Ltd",  35000, 2000, None, 10500, 2625, None, None, None),
    ]

    for row_i, (name, mktcap, netdebt, _, rev, ebitda, _, _, _) in enumerate(comps_data, start=4):
        col_vals = [
            name,
            mktcap,
            netdebt,
            f"=B{row_i}+C{row_i}",      # EV
            rev,
            ebitda,
            f"=D{row_i}/E{row_i}",      # EV/Rev
            f"=D{row_i}/F{row_i}",      # EV/EBITDA
            None,                        # P/E (manual)
        ]
        for col_i, val in enumerate(col_vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border()
            if isinstance(val, (int, float)):
                cell.font = input_font()
                cell.number_format = "#,##0"
                cell.alignment = right()
            elif isinstance(val, str) and val.startswith("="):
                cell.font = body_font(color=FORMULA_BLK)
                cell.number_format = "0.0x"
                cell.alignment = right()
            else:
                cell.font = body_font()

    # Summary rows
    summary_start = 4 + len(comps_data) + 1
    for label, func in [("Mean", "AVERAGE"), ("Median", "MEDIAN"), ("25th Pct", "PERCENTILE"), ("75th Pct", "PERCENTILE")]:
        r = summary_start
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=9, bold=True)
        for col_i in [7, 8]:
            col_letter = get_column_letter(col_i)
            if func == "PERCENTILE":
                pct = 0.25 if label == "25th Pct" else 0.75
                formula = f"=PERCENTILE({col_letter}4:{col_letter}{3+len(comps_data)},{pct})"
            else:
                formula = f"={func}({col_letter}4:{col_letter}{3+len(comps_data)})"
            cell = ws.cell(row=r, column=col_i, value=formula)
            cell.font = Font(name="Arial", size=9, bold=True, color=FORMULA_BLK)
            cell.number_format = "0.0x"
            cell.alignment = right()
            cell.border = thin_border()
        summary_start += 1


# ── Main Build Function ────────────────────────────────────────────────────────

def build_excel_model(output_path: str):
    wb = Workbook()

    # Rename default sheet
    ws1 = wb.active
    ws1.title = "Assumptions"
    build_assumptions(ws1)

    ws2 = wb.create_sheet("Income Projection")
    build_income_projection(ws2)

    ws3 = wb.create_sheet("DCF")
    build_dcf(ws3)

    ws4 = wb.create_sheet("Sensitivity")
    build_sensitivity(ws4)

    ws5 = wb.create_sheet("Comps")
    build_comps(ws5)

    # Tab colors
    ws1.sheet_properties.tabColor = DARK_NAVY
    ws2.sheet_properties.tabColor = MID_BLUE
    ws3.sheet_properties.tabColor = MID_BLUE
    ws4.sheet_properties.tabColor = "E74C3C"
    ws5.sheet_properties.tabColor = "2ECC71"

    wb.save(output_path)
    print(f"✅ Excel model saved → {output_path}")


if __name__ == "__main__":
    os.makedirs("excel", exist_ok=True)
    build_excel_model("excel/EquityValuationModel.xlsx")
